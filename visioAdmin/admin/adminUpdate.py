from typing import Dict
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import fields
from visioAdmin.admin.adminParam import AdminParam
from visioAdmin.dataModel import principale
from visioServer.models import *
from visioServer.modelStructure.dataDashboard import DataDashboard
import json
from django.core.serializers.json import DjangoJSONEncoder
import os
import shutil
from openpyxl import load_workbook
from pathlib import Path
from django.db import connection

class AdminUpdate:
  fieldNamePdv = False
  __errors = []
  __errorStatus = ""

  def __init__(self, dataDashboard):
    self.dataDashboard = dataDashboard
    AdminUpdate.__errors = []
    AdminUpdate.__errorStatus = ""
    if not AdminUpdate.fieldNamePdv:
      AdminUpdate.fieldNamePdv  = json.loads(os.getenv('FIELD_PDV_BASE_PRETTY'))
      AdminUpdate.fieldNameSales  = json.loads(os.getenv('FIELD_PDV_SALE_BASE_PRETTY'))
      AdminUpdate.titleTarget = json.loads(os.getenv('TITLE_TARGET'))
      AdminUpdate.dataSales = json.loads(os.getenv('DATA_SALES'))
      AdminUpdate.fieldXlsxRef = json.loads(os.getenv('DICO_FIELD_XLSX_REF'))
      AdminUpdate.fieldXlsxVol = json.loads(os.getenv('DICO_FIELD_XLSX_VOL'))
      AdminUpdate.fieldXlsxDbRef = json.loads(os.getenv('FIELD_PDV_XLSX_BASE'))
      AdminUpdate.pathRef = os.getenv('PATH_FILE_REF')
      AdminUpdate.pathVol = os.getenv('PATH_FILE_VOL')
      AdminUpdate.sheetNameRef = os.getenv('SHEET_NAME_REF')
      AdminUpdate.sheetNameVol = json.loads(os.getenv('SHEET_NAME_VOL'))
      AdminUpdate.keyFieldRef = os.getenv('KEY_FIELD_REF')
      AdminUpdate.keyFieldVol = os.getenv('KEY_FIELD_VOL')

      AdminUpdate.xlsxData = None
      AdminUpdate.creationElement = None
      AdminUpdate.replacedAgent = None
      AdminUpdate.response = None

  @property
  def getErrors(self):
    if self.__errors:
      return {"error":True, "title":self.__errorStatus, "content":"\r\n".join(self.__errors)}
    return False

  @classmethod
  def setError(cls, title, content):
    if title == "error":
      cls.__errorStatus = "Erreur"
    elif title == "warning" and cls.__errorStatus != "error":
      cls.__errorStatus = "Attention"
    elif title == "info" and not cls.__errorStatus:
      cls.__errorStatus = "Information"
    cls.__errors.append(content)


  def visualizeTable(self, kpi, table):
    if table == "Both":
      return self.__visualizeBoth(kpi)
    else:
      return self.__visualize(kpi, table)

  def __visualize(self, kpi, table):
    file = self.__findFileForVisualize(kpi, table)
    titles = self.fieldNamePdv if kpi == "Ref" else self.fieldNameSales
    if not os.path.isfile(f"./visioAdmin/dataFile/Json/{file}"):
      self.__createJson(kpi)
    with open(f"./visioAdmin/dataFile/Json/{file}") as jsonFile:
      pdvsToExport = json.load(jsonFile)
    return {'titles':list(titles.values()), 'values':pdvsToExport}

  def __findFileForVisualize(self, kpi, table):
    if kpi == "Ref":
      return "refSave.json" if table == "Saved" else "ref.json"
    return "volSave.json" if table == "Saved" else "vol.json"

  def __visualizeBoth(self, kpi):
    data, pdvIdCode, index, pdvsToExport, newPdv = {}, {}, 0, [], []
    fieldsName = list(self.fieldNamePdv.values()) if kpi == "Ref" else list(self.fieldNameSales.values())
    indexCode, indexClosedAt = fieldsName.index("PDV code"), fieldsName.index("Fermé le")
    dictFile = {"Current":"ref.json", "Saved":"refSave.json"} if kpi == "Ref" else {"Current":"vol.json", "Saved":"volSave.json"}
    for key, file in dictFile.items():
      self.__fillupDataForBoth(key, data, kpi, file)
    for line in data["Current"]:
      pdvIdCode[line[indexCode]] = index 
      index += 1
    for index in range(len(data["Saved"])):
      line = data["Saved"][index]
      code = line[indexCode]
      indexEquiv = pdvIdCode[code] if code in pdvIdCode else None
      if indexEquiv:
        lineEquiv = data["Current"][indexEquiv]
        lineExported = []
        for indexField in range(len(line)):
          if line[indexField] == lineEquiv[indexField]:
            lineExported.append(line[indexField])
          else:
            lineExported.append(str(line[indexField])+'<br><span class="currentvalue">'+str(lineEquiv[indexField])+'</span>')
        pdvsToExport.append(lineExported)
      else:
        line[indexClosedAt] = "Nouveau"
        pdvsToExport.append(line)
        newPdv.append(line[indexCode])
    return {'titles':fieldsName, 'values':pdvsToExport, 'new':newPdv}


  def __fillupDataForBoth(self, key, data, kpi, file):
    if not os.path.isfile(f"./visioAdmin/dataFile/Json/{file}"):
      self.__createJson(kpi)
    with open(f"./visioAdmin/dataFile/Json/{file}") as jsonFile:
      data[key] = json.load(jsonFile)

# UpdateFile
 # import Ref
  def updateFile (self, fileNature):
    if fileNature == "Ref":
      return self.__updateFileRef()
    if fileNature == "Vol":
      return self.__updateFileVol()

  def __updateFileRef(self):
    dictSheet = self.__updateLoad("Ref")
    AdminUpdate.xlsxData = dictSheet["Ref"] if dictSheet else False
    if AdminUpdate.xlsxData:
      title = self.__findTitlesRef()
      if not title: return False
      data = self.__findValuesRef(list(title.values()), title[self.keyFieldRef])
      self.__copyCurrentToSave()
      AdminUpdate.replacedAgent = self.__updateBaseSave(title, data)
      listAgent = self.__createQueryNewAgent()
      AdminUpdate.replacedAgent = listAgent
      AdminUpdate.xlsxData = {"title":list(title.keys()), "data":data}
      if listAgent:
        return {"warningAgent":listAgent}
      return self.updateRefWithAgent()

  def __copyCurrentToSave(self):
    listTable = ["visit","sales","pdv","agentFinitions","agent","dep","drv","bassin","ville","segmentCommercial","segmentMarketing","site","sousEnsemble","ensemble","enseigne"]
    with connection.cursor() as cursor:
      protectedTable = [{"field":None, "values":None, "name":"target"}, {"field":None, "values":None, "name":"targetlevel"}]
      for dictTable in protectedTable:
        cursor.execute(f"SELECT * FROM `visioServer_{dictTable['name']}`;")
        dictTable["values"] = [line for line in cursor.fetchall()]
        cursor.execute(f"SHOW FIELDS FROM `visioServer_{dictTable['name']}`;")
        dictTable["field"] = [line[0] for line in cursor.fetchall()]
        cursor.execute(f"DELETE FROM `visioServer_{dictTable['name']}`")
        cursor.execute(f"ALTER TABLE `visioServer_{dictTable['name']}` AUTO_INCREMENT=1;")

      for table in listTable:
        tableName = "visioServer_" + table.lower()
        cursor.execute(f'DELETE FROM `{tableName}save`')
        cursor.execute(f"ALTER TABLE {tableName}save AUTO_INCREMENT=1;")
      listTable.reverse()
      for table in listTable:
        tableName = "visioServer_" + table.lower()
        cursor.execute(f"SELECT * FROM `{tableName}`;")
        tableValues = [line for line in cursor.fetchall()]
        cursor.execute(f'SHOW FIELDS FROM {tableName}')
        fields = [line[0] for line in cursor.fetchall()]
        listVariable = ['%s'] * len(fields)
        strVariable = "(" + ", ".join(listVariable) + ")"
        listFields = "`,`".join(fields)
        query = f'INSERT INTO {tableName}save(`{listFields}`) VALUES {strVariable};'
        for line in tableValues:
          cursor.execute(query, line)
      for dictTable in protectedTable:
        listVariable = ['%s'] * len(dictTable["field"])
        strVariable = "(" + ", ".join(listVariable) + ")"
        listFields = "`,`".join(dictTable["field"])
        query = f'INSERT INTO visioServer_{dictTable["name"]}(`{listFields}`) VALUES {strVariable};'
        for line in dictTable["values"]:
          cursor.execute(query, line)
      
      
  def __updateFileVol(self):
    dictSheet = self.__updateLoad("Vol")
    if dictSheet:
      AdminUpdate.xlsxData = dictSheet
      titles = self.__findTitlesVol()
      if not titles: return False
      data = self.__findValuesVol(titles, titles["siniat"][self.keyFieldVol])
      AdminUpdate.xlsxData = {key:{"titles":value.keys(), "data":data[key]} for key, value in titles.items()}
      self.__saveDataVol()
    return False

  def __updateLoad(self, fileNature):
    fileName = DataAdmin.objects.get(currentBase=False).fileNameRef if fileNature == "Ref" else DataAdmin.objects.get(currentBase=False).fileNameVol
    path = self.pathRef if fileNature == "Ref" else self.pathVol
    pathFile = Path.cwd() / f"{path}{fileName}"
    dictSheet = {}
    if pathFile.exists():
      for key, sheetName in {"Ref":self.sheetNameRef}.items() if fileNature == "Ref" else self.sheetNameVol.items():
        try:
          xlsxworkbook = load_workbook(pathFile, data_only=True)
          dictSheet[key] = xlsxworkbook[sheetName]
        except:
          self.setError("error", f"L'onglet {sheetName} n'existe pas.")
          return False
      return dictSheet
    self.setError("error", f"Le fichier {fileName} n'est pas un fichier xlsx conforme.")
    return False

  def __findTitlesRef(self):
    column, dictField = 1, {}
    value = AdminUpdate.xlsxData.cell(row=1, column=column).value
    if value == self.keyFieldRef:
      while value:
        if value in self.fieldXlsxRef:
          dictField[value] = column
        column += 1
        value = AdminUpdate.xlsxData.cell(row=1, column=column).value
      for title in self.fieldXlsxRef:
        if not title in dictField:
          self.setError("error", f"Le champ {title} n'est pas présent dans les titres colonnes.")
          return False
      return dictField
    else:
      self.setError("error", f"Le champ {self.keyFieldRef} n'est pas présent en début de fichier.")
      return False

  def __findTitlesVol(self):
    dictField = {}
    for key, sheet in AdminUpdate.xlsxData.items():
      dictField[key]={}
      column = 1
      row = 2 if key == "siniat" else 1
      value = sheet.cell(row=row, column=column).value
      if value == self.keyFieldVol:
        while value:
          if value in self.fieldXlsxVol[key]:
            dictField[key][value] = column
          column += 1
          value = sheet.cell(row=row, column=column).value
        for title in self.fieldXlsxVol[key]:
          if not title in dictField[key]:
            self.setError("error", f"Le champ {title} n'est pas présent dans les titres colonnes pour les valeurs de {key}.")
            return False
      else:
        self.setError("error", f"Le champ {self.keyFieldVol} n'est pas présent dans les titres colonnes.")
        return False
    return dictField
  
  def __findValuesRef(self, titleColumns, codeColumn):
    row, data = 2, []
    while AdminUpdate.xlsxData.cell(row=row, column=codeColumn).value:
      newLine = [AdminUpdate.xlsxData.cell(row=row, column=column).value for column in titleColumns]
      data.append(newLine)
      row += 1
    return data

  def __findValuesVol(self, titleColumns, codeColumn):
    data = {key:[] for key in titleColumns.keys()}
    for key, title in titleColumns.items():
      row = 3 if key == "siniat" else 2
      while AdminUpdate.xlsxData[key].cell(row=row, column=codeColumn).value:
        newLine = [AdminUpdate.xlsxData[key].cell(row=row, column=column).value for column in title.values()]
        data[key].append(newLine)
        row += 1
    return data

  def __updateBaseSave(self, title, data):
    listId, listNewAGent = self.__buildInverseStructure(), {"new":{}, "existing":set()}
    foreignField = {field.name:field for field in PdvSave._meta.fields if isinstance(field, models.ForeignKey)}
    indexNbVisit, now = list(title.keys()).index("Nb visites F"), timezone.now()
    visitMonth = now.month - 1 if now.month != 1 else 12
    visitYear = now.year if visitMonth != 12 else now.year - 1
    dateVisit = date(year=visitYear, month=visitMonth, day=1)
    for line in data:
      pdv = self.__pdvAlreadyExists(line, listId)
      if pdv:
        indexLine = 0
        for xlsxTitle in title.keys():
          self.__updatePdv(pdv, line[indexLine], self.fieldXlsxDbRef[xlsxTitle], listId, foreignField, listNewAGent)
          indexLine += 1
        pdv.save()
      else:
        pdv = self.__createNewPdv(line, listId, list(title.keys()), listNewAGent)
      self.__computeNbVisit(pdv, line[indexNbVisit], dateVisit)
    return listNewAGent

  def __computeNbVisit(self, pdv, value, dateVisit):
    if value and isinstance(value, int):
      oldVisitList = VisitSave.objects.filter(pdv=pdv)
      for oldVisit in oldVisitList:
        if oldVisit.date.year == dateVisit.year:
          oldVisit.delete()
      VisitSave.objects.create(date=dateVisit, nbVisit=value, pdv=pdv)

  def __buildInverseStructure (self):
    """Donne pour les différents modèles la valeur clé : le name comme clé et l'Id comme valeur"""
    listId = {"pdvCode":{pdv.code:pdv.id for pdv in Pdv.objects.filter(currentYear=True)}}
    for model in [Ville, Dep, Agent, Drv, SegmentMarketing, Enseigne, Ensemble, SousEnsemble, Site, Bassin, SegmentCommercial]:
      nameM = model._meta.object_name
      listObject = model.objects.filter(currentYear=True) if "currentYear" in [field.name for field in model._meta.fields] else model.objects.all()
      listId[nameM[0].lower() + nameM[1:]] = {objectM.name:objectM.id for objectM in listObject}
    return listId

  def __pdvAlreadyExists(self, line, listId):
    IndexCodeOld = list(self.fieldXlsxDbRef.values()).index("code_old")
    IndexCodeNew = list(self.fieldXlsxDbRef.values()).index("code")
    code = line[IndexCodeOld]
    if str(code) in listId["pdvCode"]:
      pdv = PdvSave.objects.get(code=str(code), currentYear=True)
      if code != line[IndexCodeNew]:
        pdv.code = line[IndexCodeNew]
        pdvOld = PdvSave.objects.filter(code=str(code), currentYear=False)
        if pdvOld:
          pdvOld[0].code = line[IndexCodeNew]
          pdvOld[0].save()
      return pdv
    return False

  def __createNewPdv(self, line, listId, title, listNewAgent):
    indexCode = title.index("Code SAP Final")
    kwargs, indexLine = {"code":str(line[indexCode])}, 0
    foreignField = {field.name:field for field in PdvSave._meta.fields if isinstance(field, models.ForeignKey)}
    for xlsxTitle in title:
      value = line[indexLine]
      dbField = self.fieldXlsxDbRef[xlsxTitle]
      if not dbField in ["code_old", "code", "nbVisits"]:
        newValue = self.__checkForSynonym(value, dbField)
        if dbField in foreignField:
          objectFound = self.__findObjectForPdv(dbField, newValue, listId, foreignField[dbField], listNewAgent)
          if objectFound:
            kwargs[dbField] = objectFound
          elif dbField == "agent":
            kwargs[dbField] = AgentSave.objects.filter(currentYear=True)[0]
        else:
          kwargs[dbField] = newValue
      indexLine += 1
    agentFinitions = PdvSave.objects.filter(currentYear=True, dep=kwargs["dep"]).first().agentFinitions
    kwargs["agentFinitions"] = agentFinitions
    newPdv = PdvSave.objects.create(**kwargs)
    newPdv.idF = newPdv.id
    newPdv.save()
    return newPdv

  def __updatePdv(self, pdv, value, dbField, listId, foreignField, listNewAgent):
    if not dbField in ["code_old", "code", "nbVisits"]:
      newValue = self.__checkForSynonym(value, dbField)
      if dbField in foreignField:
        objectFound = self.__findObjectForPdv(dbField, newValue, listId, foreignField[dbField], listNewAgent)
        if objectFound and objectFound != getattr(pdv, dbField):
          setattr(pdv, dbField, objectFound)
        elif dbField == "agent" and value in listNewAgent["new"]:
          listNewAgent["new"][value].append(pdv)
      else:
        oldValue = getattr(pdv, dbField)
        if oldValue != newValue:
          setattr(pdv, dbField, newValue)

  def __checkForSynonym(self, value, dbField):
    synonym = False
    if dbField in list(Synonyms.dictTable.keys()):
      synonym = Synonyms.getValue(dbField, value)
    if dbField == "pointFeu":
      return True if value == "O" else False
    if not value:
      if dbField == "segmentCommercial": return "non segmenté"
    return synonym if synonym else value

  def __findObjectForPdv(self, dbField, value, listId, objectField, listNewAgent):
    if dbField == "agent":
      return self.__checkNewAgent(value, listId["agent"], listNewAgent)
    classObject = objectField.remote_field.model
    listField = [field.name for field in classObject._meta.fields]
    if value in listId[dbField]:
      return classObject.objects.get(id=listId[dbField][value])
    kwargs = {"name":value}
    if "currentYear" in listField:
      kwargs["currentYear"] = True
    objectCreated = classObject.objects.filter(**kwargs)
    if objectCreated:
      return objectCreated[0]
    objectCreated = classObject.objects.create(**kwargs)
    if "idF" in listField:
      setattr(classObject, "idF", objectCreated.id)
    return objectCreated

  def __checkNewAgent(self, value, listIdAgent, listNewAgent):
    if value in listNewAgent["new"]:
      return False
    if value in listIdAgent:
      listNewAgent["existing"].add(value)
      return AgentSave.objects.get(id=listIdAgent[value])
    listNewAgent["new"][value] = []
    return False

  def __createQueryNewAgent(self):
    listAgent = []
    listReplaced = {key:[pdv.agent.name for pdv in value] for key, value in AdminUpdate.replacedAgent["new"].items()}
    for NewName, listOldName in listReplaced.items():
      setOldName, oldNameFrequency = set(listOldName), {}
      for oldName in setOldName:
        oldNameFrequency[oldName] = listOldName.count(oldName)
        oldNameFrequencySorted = [oldName for oldName, _ in sorted(oldNameFrequency.items(), key=lambda item: -item[1])]
      if oldNameFrequencySorted[0] not in AdminUpdate.replacedAgent["existing"]:
        listAgent.append({"newName":NewName, "oldName":oldNameFrequencySorted[0]})
    return listAgent

  def updateRefWithAgent(self, getDict={}):
    pdvList = self.__updateRefWithAgent(getDict)
    self.__closePdv(pdvList)
    self.__createJson("Ref")
    shutil.copy2('./visioAdmin/dataFile/Json/vol.json', './visioAdmin/dataFile/Json/volSave.json')
    return principale.loadInit()

  def __updateRefWithAgent(self, getDict):
    dictReplacedAgent = {oldName:value[0] for oldName, value in getDict.items() if not oldName in ["action", "csrfmiddlewaretoken"]}
    pdvList = {pdv.code:pdv for pdv in PdvSave.objects.filter(currentYear=True)}
    agentListReplace = [agentReplacing for agentReplacing, status in dictReplacedAgent.items() if status == "replace"]
    agentList = {agent["newName"]:AgentSave.objects.get(name=agent["oldName"], currentYear=True) for agent in AdminUpdate.replacedAgent if agent["newName"] in agentListReplace}
    for newName, agent in agentList.items():
      agent.name = newName
      agent.save()
    indexCode = AdminUpdate.xlsxData["title"].index("Code SAP Final")
    indexAgent = AdminUpdate.xlsxData["title"].index("Libellé Agent SAP")
    for line in AdminUpdate.xlsxData["data"]:
      code = str(line[indexCode])
      agentName = line[indexAgent]
      pdv = pdvList[code] if code in pdvList else False
      if pdv and pdv.agent.name != agentName:
        newAgent = AgentSave.objects.filter(name=agentName, currentYear=True)
        if newAgent:
          newAgent = newAgent[0]
        else:
          newAgent = AgentSave.objects.create(name=agentName, currentYear=True)
          newAgent.idF = newAgent.id
          newAgent.save()
        pdv.agent = newAgent
        pdv.save()
    return pdvList

  def __closePdv(self, pdvList):
    now = timezone.now()
    indexCode = AdminUpdate.xlsxData["title"].index("Code SAP Final")
    listCode = [line[indexCode] for line in AdminUpdate.xlsxData["data"]]
    for code, pdv in pdvList.items():
      if code in listCode:
        pdv.closedAt = None
        pdv.avaliable = True
      else:
        if pdv.closedAt and pdv.closedAt.year < now.year:
          pdv.available = False
        else:
          pdv.closedAt = now
      pdv.save()


  def __saveDataVol(self):
    self.__deleteSiniatVolume()
    self.__importSiniatVolume()
    self.__createJson("Vol")

  def __createJson(self, kpi):
    if kpi == "Ref":
      self.__createRefSaveJson()
      self.__createRefJson()
    else:
      self.__createVolJson()

  def __createRefSaveJson(self):
    pdvs = PdvSave.objects.filter(currentYear=True)
    listName = PdvSave.listFields()
    fieldsObject = [name for name in listName if getattr(Pdv, name, False) and (isinstance(Pdv._meta.get_field(name), models.ForeignKey))]
    pdvsToExport = [self.__computePdvSaveJson(pdv, fieldsObject) for pdv in pdvs]
    with open("./visioAdmin/dataFile/Json/refSave.json", 'w') as jsonFile:
      json.dump(pdvsToExport, jsonFile, indent = 3)

  def __createRefJson(self):
    pdvs = getattr(self.dataDashboard, "__pdvs")
    indexes = Pdv.listIndexes()
    listFields = Pdv.listFields()
    pdvsToExport = [self.__computePdvJson(line, listFields, indexes, self.fieldNamePdv) for line in pdvs.values()]
    with open("./visioAdmin/dataFile/Json/ref.json", 'w') as jsonFile:
      json.dump(pdvsToExport, jsonFile, indent = 3)

  def __computePdvSaveJson(self, pdv, fieldsObject):
    lineFormated = []
    for  field in self.fieldNamePdv.keys():
      fieldVal = getattr(pdv, field)
      value = ""
      if field == "closedAt":
        value = fieldVal.strftime("%m/%d/%Y") if pdv.closedAt else ""
      elif isinstance(fieldVal, bool):
        value = "Oui" if fieldVal else "Non"
      elif field == "bassin":
        value = fieldVal.name.replace('Négoce_', "")
      elif field in fieldsObject:
        value = fieldVal.name
      else:
        value = fieldVal
      lineFormated.append(value)
    return lineFormated

  def __computePdvJson(self, line, listFields, indexes, fieldNamePdv):
    lineFormated = []
    for index in range(len(line)):
      if listFields[index] in fieldNamePdv:
        if index in indexes:
          dictData = getattr(DataDashboard, f"__{listFields[index]}", False)
          if dictData:
            value = dictData[line[index]]
            if isinstance(value, list):
              value = value[0]
            lineFormated.append(value)
        elif isinstance(line[index], bool):
          lineFormated.append("Oui" if line[index] else "Non")
        elif listFields[index] == "closedAt":
            lineFormated.append(line[index][:10] if line[index] else '')
        else:
          lineFormated.append(line[index])
    return lineFormated

  # import Vol
  def __deleteSiniatVolume(self):
    for industry in ["Siniat", "Prégy", "Salsi"]:
      industryObj = Industry.objects.get(name=industry)
      SalesSave.objects.filter(industry=industryObj, currentYear=True).delete()

  def __importSiniatVolume(self):
    siniat = Industry.objects.get(name="Siniat")
    pregy = Industry.objects.get(name="Prégy")
    salsi = Industry.objects.get(name="Salsi")
    plaque =  Product.objects.get(name="plaque")
    cloison =  Product.objects.get(name="cloison")
    doublage =  Product.objects.get(name="doublage")
    enduit = Product.objects.get(name="enduit")
    listTitlesSiniat = list(self.xlsxData["siniat"]["titles"])
    listTitlesSalsi = list(self.xlsxData["salsi"]["titles"])
    indexPdv = listTitlesSiniat.index("code pdv")
    indexPlaque = listTitlesSiniat.index("Plaques")
    indexCloison = listTitlesSiniat.index("Cloisons")
    indexDoublage = listTitlesSiniat.index("Doublages")
    indexPregy = listTitlesSiniat.index("Enduit Prégy")
    indexSalsi = listTitlesSalsi.index("Enduit Joint Salsi")
    dictSiniat = [{"ind":siniat, "prod":plaque, "iVol":indexPlaque}, {"ind":siniat, "prod":cloison, "iVol":indexCloison}, {"ind":siniat, "prod":doublage, "iVol":indexDoublage}, {"ind":pregy, "prod":enduit, "iVol":indexPregy}]
    for data in self.xlsxData["siniat"]["data"]:
      for dictData in dictSiniat:
        if data[dictData["iVol"]]:
          pdv = PdvSave.objects.filter(code=data[indexPdv], currentYear=True)
          if pdv:
            SalesSave.objects.create(date=None, pdv=pdv[0], industry=dictData["ind"], product=dictData["prod"],volume=data[dictData["iVol"]], currentYear=True)
    for data in self.xlsxData["salsi"]["data"]:
      if data[indexSalsi]:
        pdv = PdvSave.objects.filter(code=data[indexPdv], currentYear=True)
        if pdv:
          SalesSave.objects.create(date=None, pdv=pdv[0], industry=salsi, product=enduit,volume=data[indexSalsi], currentYear=True)

  def __createVolJson(self):
    for nature in ["Current", "Saved"]:
      fileRef, dictSales = "vol.json" if nature == "Current" else "volSave.json", {}
      listSales = [objectSales for objectSales in Sales.objects.filter(currentYear=True)] if nature == "Current" else [objectSales for objectSales in SalesSave.objects.all()]
      for sale in listSales:
        self.__computeVolJson(sale, dictSales)
      with open(f"./visioAdmin/dataFile/Json/{fileRef}", 'w') as jsonFile:
        json.dump(list(dictSales.values()), jsonFile, indent = 3)
    

  def __computeVolJson(self, sales, dictSales):
    if sales.industry.name in ["Siniat", "Prégy", "Salsi"] and sales.product.name != "mortier":
      if not sales.pdv.code in dictSales:
        lineFormated = [sales.pdv.code, sales.pdv.name]
        lineFormated.append("Non" if sales.pdv.sale else "Oui")
        lineFormated.append("Non" if sales.pdv.redistributed else "Oui")
        lineFormated.append("Non" if sales.pdv.redistributedFinitions else "Oui")
        lineFormated.append("Oui" if sales.pdv.onlySiniat else "Non")
        lineFormated.append(sales.pdv.closedAt.strftime("%m/%d/%Y") if sales.pdv.closedAt else "")
        lineFormated += [0, 0, 0, 0, 0]
        dictSales[sales.pdv.code] = lineFormated
      lineFormated = dictSales[sales.pdv.code]
      index = False
      if sales.industry.name == "Siniat": 
        if sales.product.name == "plaque": index = 7
        if sales.product.name == "cloison": index = 8
        if sales.product.name == "doublage": index = 9
      if sales.industry.name == "Prégy" and sales.product.name == "enduit": index = 10
      if sales.industry.name == "Salsi" and sales.product.name == "enduit": index = 11
      if index:
        dictSales[sales.pdv.code][index] = '{:,}'.format(int(sales.volume)).replace(',', ' ')

  # switch Base

  def switchBase(self):
    listTable = ["visit","sales","pdv","agentFinitions","agent","dep","drv","bassin","ville","segmentCommercial","segmentMarketing","site","sousEnsemble","ensemble","enseigne"]
    extention = {"current":"", "save":"save", "temp":"temp"}
    listRename = [
      {"from":extention["current"], "to":extention["temp"]},
      {"from":extention["save"], "to":extention["current"]},
      {"from":extention["temp"], "to":extention["save"]}
      ]
    for rename in listRename:
      self.__renameTable(listTable, rename)
    dataAdmin = DataAdmin.objects.get(currentBase=False)
    dataAdminCurrent = DataAdmin.objects.get(currentBase=True)
    dataAdmin.currentBase = True
    dataAdminCurrent.currentBase = False
    dataAdmin.save()
    dataAdminCurrent.save()
    renameActions = [
      {"from":"ref.json", "to":"refTemp.json"},
      {"from":"refSave.json", "to":"ref.json"},
      {"from":"refTemp.json", "to":"refSave.json"},
      {"from":"vol.json", "to":"volTemp.json"},
      {"from":"volSave.json", "to":"vol.json"},
      {"from":"volTemp.json", "to":"volSave.json"},
      ]
    for action in renameActions:
      os.rename(f'./visioAdmin/dataFile/Json/{action["from"]}', f'./visioAdmin/dataFile/Json/{action["to"]}')
    ParamVisio.setValue("referentielVersion", dataAdmin.getVersion)

  def __renameTable(self, listTable, rename):
    with connection.cursor() as cursor:
      for table in listTable:
        tableNameFrom = "visioServer_" + table.lower() + rename["from"]
        tableNameTo = "visioServer_" + table.lower() + rename["to"]
        cursor.execute(f'RENAME TABLE `{tableNameFrom}` TO `{tableNameTo}`')

    









