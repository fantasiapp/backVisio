import openpyxl
from openpyxl.styles import Font
from django.utils import timezone
import os
import json
from pathlib import Path
from django.http import HttpResponse
from visioServer.models import Target, PdvSave

import os.path
from os import path

class AdminConsult:
  pathSave = os.getenv('PATH_FILE_SEND')
  print()
  dictFileName =  json.loads(os.getenv('DICO_FILE_SEND'))
  extention = "xlsx"

  def __init__(self, adminUpdate, adminParam):
    self.adminUpdate = adminUpdate
    self.adminParam = adminParam
    self.dataDashboard = adminUpdate.dataDashboard

  def buildExcelFile(self, nature):
    data = self.__findData(nature)
    return self.__createExcelFile(data, nature)

  def __findData(self, nature):
    if nature == "currentBase":
      dataRef = self.adminUpdate.visualizeTable("Ref", "Current")
      dataVol = self.adminUpdate.visualizeTable("Vol", "Current")
      return {"Référentiel de la base courante":dataRef, "Volume de la base courante":dataVol}
    if nature == "connection":
      dataConnection = self.adminParam.paramAccountInit()
      dataConnection["titles"] = list(dataConnection["titles"].keys())
      dataConnection["values"] = list(dataConnection["values"].values())
      return {"Rapport des connexions":dataConnection}
    if nature == "target":
      dataRef = self.visualizeTargetTable("Ref")
      dataTarget = self.visualizeTargetTable("Target")
      dataTarget["titles"] = list(dataRef["titles"].keys())
      dataTarget["titles"] = list(dataRef["titles"].keys())
      return {"Modifications demandées":dataRef, "Ciblage":dataTarget}



  def __createExcelFile(self, data, nature):
    file, flagStart = openpyxl.Workbook(), True
    for sheetName, data in data.items():
      if flagStart:
        sheet = file.active
        sheet.title = sheetName
        flagStart = False
      else:
        sheet = file.create_sheet(sheetName)
      self.__createTitle(sheet, data["titles"])
      self.__createLine(sheet, data["values"])
    return self.__saveExcelFile(file, nature)

  def __createTitle(self, sheet, titles):
    color = openpyxl.styles.colors.Color(rgb='0B64A0')
    fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=color)
    fontColor = Font(color="FFFFFF", bold=True)
    nbRow, nbCol = 1, 1
    for title in titles:
      cell = sheet.cell(row=nbRow, column=nbCol)
      cell.value = title
      nbCol += 1
      cell.fill = fill
      cell.font = fontColor

  def __createLine(self, sheet, values):
    nbRow = 2
    for line in values:
      nbCol = 1
      for value in line:
        cell = sheet.cell(row=nbRow, column=nbCol)
        cell.value = value
        nbCol += 1
      nbRow += 1

  def __saveExcelFile(self, file, nature):
    now = timezone.now()
    name = f"{self.dictFileName[nature]}_{now.strftime('%d_%m_%y_%Hh%Mmn_%Ss')}.{self.extention}"
    fileName = f"{Path.cwd()}/{self.pathSave}{name}"
    file.save(fileName)
    # content = FileWrapper(fileName)
    with open(fileName, 'rb') as file:
      response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
      response['Content-Length'] = os.path.getsize(fileName)
      response['Content-Disposition'] = 'inline; filename=%s' % name
      return response


  def visualizeTargetTable(self, table):
    if table == "Ref":
      response = self.adminParam.buildValidate(target=True)
      newValues = []
      for key, dictValues in response["values"].items():
        for line in dictValues.values():
          line.insert(5, key)
          newValues.append(line)
      response["values"] = newValues
      return response
    return self.visualizeTargetRef()

  def visualizeTargetRef(self):
    titles = {"Drv":10, "Agent":15, "Pdv Code":7, "Date":8, "Pdv":28, "Cible P2CD":12, "Cible Finition":12, "Feu": 8}
    values = [self.__buildTargetLine(target) for target in Target.objects.all() if self.__testValidateLine(target.pdv, target)]
    return {"titles":titles, "values":values}

  def __buildTargetLine(self, target):
    pdvId = target.pdv.id
    pdv = PdvSave.objects.get(id=pdvId)
    finitions = "Oui" if target.targetFinitions else "Non"
    if target.greenLight == "g":
      feu = "vert"
    elif target.greenLight == "o":
      feu = "orange"
    elif target.greenLight == "r":
      feu = "rouge"
    return [pdv.drv.name, pdv.agent.name, pdv.code, target.date.strftime('%Y-%m-%d'), pdv.name, "{:.2f}".format(target.targetP2CD), finitions, feu]

  def __testValidateLine(self, pdv, target):
    return pdv.available and pdv.sale and pdv.redistributed and pdv.currentYear and (target.targetP2CD or target.targetFinitions)



